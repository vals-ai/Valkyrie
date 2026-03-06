from __future__ import annotations

import asyncio
import json
import logging
import os
import signal
from pathlib import Path
from typing import Any

from model_library.agent.tool import Tool, ToolOutput

_MACRO_DIR = Path.home() / ".config/libreoffice/4/user/basic/Standard"
_MACRO_FILE = _MACRO_DIR / "Module1.xba"

_MACRO_CONTENT = """\
<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
Sub RecalculateAndSave
    Dim oDoc As Object
    Dim args(0) As New com.sun.star.beans.PropertyValue
    Dim sURL As String

    sURL = ConvertToURL(Command())
    args(0).Name = "Hidden"
    args(0).Value = True

    oDoc = StarDesktop.loadComponentFromURL(sURL, "_blank", 0, args())
    If IsNull(oDoc) Or IsEmpty(oDoc) Then
        MsgBox "Failed to open: " &amp; sURL
        StarDesktop.terminate()
    End If

    oDoc.calculateAll()
    oDoc.store()
    oDoc.close(True)
    StarDesktop.terminate()
End Sub
</script:module>
"""

_EXCEL_ERRORS = {"#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}
_MAX_ERRORS_PER_TYPE = 20
_SOFFICE_TIMEOUT = 60


class RecalculateExcelTool(Tool):
    name = "recalculate_excel"
    description = (
        "Recalculate all formulas in the Excel workbook using LibreOffice. "
        "Returns error summary and formula count. After calling this, "
        "read specific cells via openpyxl with data_only=True to inspect computed values."
    )
    parameters: dict[str, Any] = {}
    required: list[str] = []

    def __init__(self, *, file_path: str, **kwargs: Any):
        super().__init__(**kwargs)
        self._file_path = file_path

    def _install_macro(self) -> None:
        if _MACRO_FILE.exists():
            return
        _MACRO_DIR.mkdir(parents=True, exist_ok=True)
        _MACRO_FILE.write_text(_MACRO_CONTENT)

    async def _run_soffice(self, logger: logging.Logger) -> tuple[int, str, str]:
        cmd = (
            f"soffice --headless --norestore "
            f"'vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application' "
            f"'{self._file_path}'"
        )
        logger.info(f"running: {cmd}")

        proc = await asyncio.create_subprocess_shell(
            cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
            start_new_session=True,
        )

        try:
            stdout_bytes, stderr_bytes = await asyncio.wait_for(
                proc.communicate(), timeout=_SOFFICE_TIMEOUT
            )
        except asyncio.TimeoutError:
            try:
                os.killpg(os.getpgid(proc.pid), signal.SIGKILL)
            except ProcessLookupError:
                pass
            await proc.wait()
            return -1, "", f"LibreOffice timed out after {_SOFFICE_TIMEOUT}s"

        stdout = stdout_bytes.decode(errors="replace")
        stderr = stderr_bytes.decode(errors="replace")
        return proc.returncode or 0, stdout, stderr

    def _scan_errors(self) -> tuple[int, dict[str, list[str]]]:
        import openpyxl

        wb = openpyxl.load_workbook(self._file_path, data_only=True)
        error_map: dict[str, list[str]] = {}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value in _EXCEL_ERRORS:
                        total_errors += 1
                        locations = error_map.setdefault(cell.value, [])
                        if len(locations) < _MAX_ERRORS_PER_TYPE:
                            locations.append(f"{sheet_name}!{cell.coordinate}")
        wb.close()
        return total_errors, error_map

    def _count_formulas(self) -> int:
        import openpyxl

        wb = openpyxl.load_workbook(self._file_path, data_only=False)
        count = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        count += 1
        wb.close()
        return count

    async def execute(
        self, args: dict[str, Any], state: dict[str, Any], logger: logging.Logger
    ) -> ToolOutput:
        path = Path(self._file_path)
        if not path.exists():
            msg = f"File not found: {self._file_path}"
            return ToolOutput(output=msg, error=msg)
        if path.suffix.lower() != ".xlsx":
            msg = f"Expected .xlsx file, got: {path.suffix}"
            return ToolOutput(output=msg, error=msg)

        self._install_macro()

        exit_code, stdout, stderr = await self._run_soffice(logger)
        if exit_code != 0:
            msg = f"LibreOffice failed (exit {exit_code}): {stderr}"
            return ToolOutput(output=msg, error=msg)

        total_errors, error_map = self._scan_errors()
        total_formulas = self._count_formulas()

        error_summary = {
            err_type: {"count": len(locs), "locations": locs}
            for err_type, locs in error_map.items()
        }

        result = {
            "status": "ok" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "total_formulas": total_formulas,
            "error_summary": error_summary,
        }

        return ToolOutput(output=json.dumps(result, indent=2))
